from __future__ import annotations

import json
from pathlib import Path
from typing import List, Dict

import numpy as np
import pandas as pd

# For dashboard exports
import matplotlib.pyplot as plt

# For Excel dashboard
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# For final memo PDF
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas


# =========================
# Paths (as per hierarchy)
# =========================
PROJECT_ROOT = Path(__file__).resolve().parents[1]

RAW_DIR = PROJECT_ROOT / "raw_layer"

DATA_DIR = PROJECT_ROOT / "data"                  # generated outputs only
ANALYSIS_DIR = PROJECT_ROOT / "analysis"           # generated analysis artifacts
DASHBOARD_DIR = PROJECT_ROOT / "dashboard"         # dashboard file
EXPORTS_DIR = DASHBOARD_DIR / "exports"            # charts/images/pdf exports
FINAL_STORY_DIR = PROJECT_ROOT / "final_story"     # final memo/deck

for d in [DATA_DIR, ANALYSIS_DIR, DASHBOARD_DIR, EXPORTS_DIR, FINAL_STORY_DIR]:
    d.mkdir(parents=True, exist_ok=True)


# =========================
# Helpers: IO
# =========================
def read_csv(name: str) -> pd.DataFrame:
    path = RAW_DIR / name
    if not path.exists():
        raise FileNotFoundError(f"Missing raw file: {path}")
    return pd.read_csv(path)


def read_json(name: str):
    path = RAW_DIR / name
    if not path.exists():
        raise FileNotFoundError(f"Missing raw file: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def write_csv(df: pd.DataFrame, path: Path) -> None:
    df.to_csv(path, index=False)
    print(f"✅ Wrote: {path}")


# =========================
# Helpers: Cleaning
# =========================
def to_snake_case(cols: List[str]) -> List[str]:
    out = []
    for c in cols:
        c2 = str(c).strip().replace(" ", "_").replace("-", "_")
        out.append(c2.lower())
    return out


def standardize_df(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = to_snake_case(df.columns)
    for c in df.columns:
        if df[c].dtype == "object":
            df[c] = df[c].astype("string").str.strip()
    return df


def safe_lower(s: pd.Series) -> pd.Series:
    return s.astype("string").str.strip().str.lower()


def parse_dt(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    return df


def dedup_keep_last(df: pd.DataFrame, key_cols: List[str], order_col: str | None = None) -> pd.DataFrame:
    df = df.copy()
    if order_col and order_col in df.columns:
        df = df.sort_values(order_col)
    return df.drop_duplicates(subset=key_cols, keep="last")


# =========================
# Helpers: Features
# =========================
def discount_pct(gross: pd.Series, discount: pd.Series) -> pd.Series:
    gross = pd.to_numeric(gross, errors="coerce").fillna(0)
    discount = pd.to_numeric(discount, errors="coerce").fillna(0)
    denom = gross.replace(0, np.nan)
    pct = (discount / denom) * 100
    return pct.fillna(0).clip(lower=0)


def week_start(dt: pd.Series) -> pd.Series:
    d = pd.to_datetime(dt, errors="coerce")
    return (d - pd.to_timedelta(d.dt.weekday, unit="D")).dt.date


def zscore_within_group(df: pd.DataFrame, group_col: str, value_col: str) -> pd.Series:
    def _z(v: pd.Series) -> pd.Series:
        v = pd.to_numeric(v, errors="coerce").fillna(0)
        if len(v) < 5:
            return pd.Series([0] * len(v), index=v.index)
        std = v.std(ddof=0)
        if std == 0 or np.isnan(std):
            return pd.Series([0] * len(v), index=v.index)
        return (v - v.mean()) / std

    return df.groupby(group_col)[value_col].transform(_z).fillna(0)


# =========================
# Risk scoring config
# =========================
WEIGHTS = {
    "high_discount_flag": 14,
    "discount_severity_score": 12,
    "coupon_repeat_user_flag": 12,
    "coupon_device_reuse_flag": 14,
    "payment_failed_attempts_score": 20,
    "new_user_flag": 8,
    "new_user_plus_coupon": 12,
    "cod_flag": 6,
    "high_rto_pincode_flag": 14,
    "pincode_reuse_score": 12,
    "device_reuse_score": 14,
    "value_outlier_flag": 8,
    "refund_history_user_flag": 12,
}

REASON_CODES = {
    "high_discount_flag": "HIGH_DISCOUNT",
    "discount_severity_score": "DISCOUNT_SEVERITY",
    "coupon_repeat_user_flag": "COUPON_REPEAT_USER",
    "coupon_device_reuse_flag": "COUPON_DEVICE_REUSE",
    "payment_failed_attempts_score": "PAYMENT_FAIL_SPIKE",
    "new_user_flag": "NEW_USER",
    "new_user_plus_coupon": "NEW_USER_COUPON",
    "cod_flag": "COD_ORDER",
    "high_rto_pincode_flag": "HIGH_RTO_PINCODE",
    "pincode_reuse_score": "PINCODE_REUSE",
    "device_reuse_score": "DEVICE_REUSE",
    "value_outlier_flag": "VALUE_OUTLIER",
    "refund_history_user_flag": "REFUND_HISTORY",
}


def clamp01(s: pd.Series) -> pd.Series:
    return s.fillna(0).clip(0, 1)


def compute_score_and_reasons(fact: pd.DataFrame) -> pd.DataFrame:
    df = fact.copy()

    for f in WEIGHTS.keys():
        if f not in df.columns:
            df[f] = 0

    contrib_cols = []
    for feat, w in WEIGHTS.items():
        v = df[feat]
        if feat.endswith("_score"):
            v = clamp01(pd.to_numeric(v, errors="coerce"))
        else:
            v = pd.to_numeric(v, errors="coerce").fillna(0).clip(0, 1)
        ccol = f"contrib__{feat}"
        df[ccol] = w * v
        contrib_cols.append(ccol)

    df["risk_score"] = df[contrib_cols].sum(axis=1).round().clip(0, 100).astype(int)
    df["risk_band"] = pd.cut(
    df["risk_score"],
    bins=[-1, 29, 49, 100],
    labels=["Low", "Medium", "High"]
    ).astype("string")

    feat_from_col = {c: c.replace("contrib__", "") for c in contrib_cols}

    def top3(row) -> List[str]:
        pairs = []
        for c in contrib_cols:
            val = row[c]
            if pd.notna(val) and val > 0:
                feat = feat_from_col[c]
                pairs.append((val, REASON_CODES.get(feat, feat)))
        pairs.sort(reverse=True, key=lambda x: x[0])
        reasons = [p[1] for p in pairs[:3]]
        reasons += [""] * (3 - len(reasons))
        return reasons

    reasons_df = df.apply(lambda r: pd.Series(top3(r)), axis=1)
    reasons_df.columns = ["reason_1", "reason_2", "reason_3"]
    df = pd.concat([df, reasons_df], axis=1)

    df.drop(columns=contrib_cols, inplace=True, errors="ignore")
    return df


def recommend_action(row: pd.Series) -> str:
    band = row.get("risk_band", "")
    reasons = {row.get("reason_1", ""), row.get("reason_2", ""), row.get("reason_3", "")}

    if band == "High":
        if ("DEVICE_REUSE" in reasons or "PINCODE_REUSE" in reasons) and ("NEW_USER_COUPON" in reasons or "NEW_USER" in reasons):
            return "HOLD"
        if "PAYMENT_FAIL_SPIKE" in reasons:
            return "CALL_VERIFICATION"
        if "HIGH_RTO_PINCODE" in reasons and "COD_ORDER" in reasons:
            return "AUTO_CANCEL"
        return "MANUAL_REVIEW"

    if band == "Medium":
        if "HIGH_DISCOUNT" in reasons or "NEW_USER_COUPON" in reasons:
            return "CALL_VERIFICATION"
        return "MANUAL_REVIEW"

    return "ALLOW"

# =========================
# Products flatten
# =========================
def flatten_products(products_json) -> pd.DataFrame:
    if isinstance(products_json, dict) and "products" in products_json:
        products_json = products_json["products"]

    rows = []
    if isinstance(products_json, list):
        for p in products_json:
            if not isinstance(p, dict):
                continue
            rows.append({
                "product_id": p.get("product_id") or p.get("id") or p.get("sku"),
                "category": p.get("category") or p.get("product_category") or "unknown",
            })
    return pd.DataFrame(rows)


# =========================
# Build Fact: 1 row per order
# =========================
def build_fact_orders_enriched(
    users: pd.DataFrame,
    sessions: pd.DataFrame,
    orders: pd.DataFrame,
    order_items: pd.DataFrame,
    payments: pd.DataFrame,
    shipments: pd.DataFrame,
    refunds: pd.DataFrame,
    coupons: pd.DataFrame,
    products: pd.DataFrame
) -> pd.DataFrame:

    users = parse_dt(standardize_df(users), ["signup_ts", "created_at"])
    sessions = parse_dt(standardize_df(sessions), ["session_ts", "created_at"])
    orders = parse_dt(standardize_df(orders), ["order_ts", "created_at"])
    order_items = standardize_df(order_items)
    payments = parse_dt(standardize_df(payments), ["payment_ts", "created_at"])
    shipments = parse_dt(standardize_df(shipments), ["shipment_ts", "created_at"])
    refunds = parse_dt(standardize_df(refunds), ["refund_ts", "created_at"])
    coupons = standardize_df(coupons)
    products = standardize_df(products)

    if "user_id" in users.columns:
        users = dedup_keep_last(users, ["user_id"], order_col="signup_ts" if "signup_ts" in users.columns else None)

    if "session_id" in sessions.columns:
        sessions = dedup_keep_last(sessions, ["session_id"], order_col="session_ts" if "session_ts" in sessions.columns else None)

    if "order_id" in orders.columns:
        orders = dedup_keep_last(orders, ["order_id"], order_col="order_ts" if "order_ts" in orders.columns else None)

    for c in ["order_id", "user_id", "session_id", "gross_amount", "discount_amount", "net_amount", "payment_method", "coupon_id"]:
        if c not in orders.columns:
            orders[c] = pd.NA

    orders["payment_method"] = safe_lower(orders["payment_method"]).fillna("unknown")
    orders["coupon_id"] = safe_lower(orders["coupon_id"]).fillna("none")

    if "shipping_city" in orders.columns:
        orders["shipping_city"] = safe_lower(orders["shipping_city"]).fillna("unknown")

    if "shipping_pincode" in orders.columns:
        orders["shipping_pincode"] = orders["shipping_pincode"].astype("string").str.strip().fillna("unknown")
    else:
        orders["shipping_pincode"] = "unknown"

    for c in ["gross_amount", "discount_amount", "net_amount"]:
        orders[c] = pd.to_numeric(orders[c], errors="coerce").fillna(0)

    orders["discount_pct"] = discount_pct(orders["gross_amount"], orders["discount_amount"])

    if "product_id" in products.columns:
        products["product_id"] = products["product_id"].astype("string").str.strip()
    if "product_id" in order_items.columns:
        order_items["product_id"] = order_items["product_id"].astype("string").str.strip()

    if "product_id" in order_items.columns and "product_id" in products.columns:
        order_items = order_items.merge(products[["product_id", "category"]].drop_duplicates(), on="product_id", how="left")

    if "qty" not in order_items.columns:
        order_items["qty"] = 1
    order_items["qty"] = pd.to_numeric(order_items["qty"], errors="coerce").fillna(1)

    items_agg = order_items.groupby("order_id", as_index=False).agg(
        item_count=("product_id", "nunique") if "product_id" in order_items.columns else ("order_id", "size"),
        total_qty=("qty", "sum"),
    )

    if "category" in order_items.columns:
        tmp = order_items.copy()
        tmp["category"] = tmp["category"].astype("string").fillna("unknown").str.lower().str.strip()
        cat_qty = tmp.groupby(["order_id", "category"], as_index=False)["qty"].sum()
        cat_qty = cat_qty.sort_values(["order_id", "qty"], ascending=[True, False])
        top_cat = cat_qty.drop_duplicates("order_id", keep="first").rename(columns={"category": "top_category"})
        items_agg = items_agg.merge(top_cat[["order_id", "top_category"]], on="order_id", how="left")
    else:
        items_agg["top_category"] = "unknown"

    if "status" in payments.columns:
        payments["status"] = safe_lower(payments["status"])
        payments["is_fail"] = payments["status"].str.contains("fail", na=False).astype(int)
        payments["is_success"] = payments["status"].str.contains("success", na=False).astype(int)
    else:
        payments["is_fail"] = 0
        payments["is_success"] = 0

    pay_agg = payments.groupby("order_id", as_index=False).agg(
    payment_attempts=("order_id", "size"),
    failed_attempts=("is_fail", "sum"),
    success_attempts=("is_success", "sum"),
    )

    pay_agg["payment_failed_attempts_score"] = pay_agg["failed_attempts"].map(
    lambda x: 0.0 if x <= 0 else 0.35 if x == 1 else 0.65 if x == 2 else 1.0
    )
    if "refund_amount" not in refunds.columns:
        refunds["refund_amount"] = 0
    refunds["refund_amount"] = pd.to_numeric(refunds["refund_amount"], errors="coerce").fillna(0)

    refund_agg = refunds.groupby("order_id", as_index=False).agg(
        refund_flag=("refund_amount", lambda x: int(x.sum() > 0)),
        refund_amount=("refund_amount", "sum"),
    )

    if "status" in shipments.columns:
        shipments["status"] = safe_lower(shipments["status"])
        shipments["rto_flag"] = shipments["status"].str.contains("rto", na=False).astype(int)
        shipments["delivered_flag"] = shipments["status"].str.contains("deliver", na=False).astype(int)
    else:
        shipments["rto_flag"] = 0
        shipments["delivered_flag"] = 0

    ship_agg = shipments.groupby("order_id", as_index=False).agg(
        rto_flag=("rto_flag", "max"),
        delivered_flag=("delivered_flag", "max"),
    )

    if "coupon_id" in coupons.columns:
        coupons["coupon_id"] = safe_lower(coupons["coupon_id"]).fillna("none")

    if "discount_pct" in coupons.columns:
        coupons["coupon_discount_pct"] = pd.to_numeric(coupons["discount_pct"], errors="coerce")
    elif "coupon_discount_pct" not in coupons.columns:
        coupons["coupon_discount_pct"] = pd.NA

    coupons_ref = coupons[["coupon_id", "coupon_discount_pct"]].drop_duplicates()

    fact = orders.copy()
    if "user_id" in users.columns:
        fact = fact.merge(users, on="user_id", how="left", suffixes=("", "_user"))
    if "session_id" in sessions.columns:
        fact = fact.merge(sessions, on="session_id", how="left", suffixes=("", "_session"))

    fact = fact.merge(items_agg, on="order_id", how="left")
    fact = fact.merge(pay_agg, on="order_id", how="left")
    fact = fact.merge(refund_agg, on="order_id", how="left")
    fact = fact.merge(ship_agg, on="order_id", how="left")
    fact = fact.merge(coupons_ref, on="coupon_id", how="left")

    for c, d in [
        ("item_count", 0), ("total_qty", 0), ("top_category", "unknown"),
        ("payment_attempts", 0), ("failed_attempts", 0), ("success_attempts", 0),
        ("payment_failed_attempts_score", 0),
        ("refund_flag", 0), ("refund_amount", 0),
        ("rto_flag", 0), ("delivered_flag", 0),
    ]:
        if c in fact.columns:
            fact[c] = fact[c].fillna(d)

    # ---- engineered signals ----
    fact["high_discount_flag"] = (fact["discount_pct"] >= 50).astype(int)
    fact["coupon_used_flag"] = (fact["coupon_id"].astype("string") != "none").astype(int)
    fact["cod_flag"] = fact["payment_method"].astype("string").str.contains("cod", na=False).astype(int)

    signup_col = "signup_ts" if "signup_ts" in fact.columns else ("created_at" if "created_at" in fact.columns else None)
    order_col = "order_ts" if "order_ts" in fact.columns else ("created_at" if "created_at" in fact.columns else None)

    if signup_col and order_col:
        days = (pd.to_datetime(fact[order_col], errors="coerce") - pd.to_datetime(fact[signup_col], errors="coerce")).dt.days
        fact["days_since_signup"] = days.fillna(9999).astype(int)
        fact["new_user_flag"] = (fact["days_since_signup"] <= 7).astype(int)
    else:
        fact["days_since_signup"] = 9999
        fact["new_user_flag"] = 0

    fact["new_user_plus_coupon"] = ((fact["new_user_flag"] == 1) & (fact["coupon_used_flag"] == 1)).astype(int)

    if "user_id" in fact.columns:
        user_ref = fact.groupby("user_id")["refund_flag"].sum().rename("prior_refunds_count").reset_index()
        fact = fact.merge(user_ref, on="user_id", how="left")
        fact["prior_refunds_count"] = fact["prior_refunds_count"].fillna(0).astype(int)
        fact["refund_history_user_flag"] = (fact["prior_refunds_count"] >= 2).astype(int)
    else:
        fact["prior_refunds_count"] = 0
        fact["refund_history_user_flag"] = 0

    if "device_id" in fact.columns:
        fact["device_id"] = fact["device_id"].astype("string").str.strip().fillna("unknown")
        device_cnt = fact.groupby("device_id")["order_id"].count().rename("device_orders_count").reset_index()
        fact = fact.merge(device_cnt, on="device_id", how="left")
        fact["device_orders_count"] = fact["device_orders_count"].fillna(1).astype(int)
        fact["device_reuse_score"] = (fact["device_orders_count"].clip(upper=20) / 20.0)
    else:
        fact["device_orders_count"] = 1
        fact["device_reuse_score"] = 0

    pin_cnt = fact.groupby("shipping_pincode")["order_id"].count().rename("pincode_orders_count").reset_index()
    fact = fact.merge(pin_cnt, on="shipping_pincode", how="left")
    fact["pincode_orders_count"] = fact["pincode_orders_count"].fillna(1).astype(int)
    fact["pincode_reuse_score"] = (fact["pincode_orders_count"].clip(upper=20) / 20.0)

    if "user_id" in fact.columns:
        user_coupon = fact[fact["coupon_used_flag"] == 1].groupby("user_id")["order_id"].count().rename("coupon_orders_user").reset_index()
        fact = fact.merge(user_coupon, on="user_id", how="left")
        fact["coupon_orders_user"] = fact["coupon_orders_user"].fillna(0).astype(int)
        fact["coupon_repeat_user_flag"] = (fact["coupon_orders_user"] >= 3).astype(int)
    else:
        fact["coupon_orders_user"] = 0
        fact["coupon_repeat_user_flag"] = 0

    if "device_id" in fact.columns:
        dev_coupon_users = (
            fact[fact["coupon_used_flag"] == 1]
            .groupby("device_id")["user_id"].nunique()
            .rename("coupon_users_per_device")
            .reset_index()
        )
        fact = fact.merge(dev_coupon_users, on="device_id", how="left")
        fact["coupon_users_per_device"] = fact["coupon_users_per_device"].fillna(0).astype(int)
        fact["coupon_device_reuse_flag"] = (fact["coupon_users_per_device"] >= 3).astype(int)
    else:
        fact["coupon_users_per_device"] = 0
        fact["coupon_device_reuse_flag"] = 0

    p = fact.groupby("shipping_pincode").agg(
        pincode_orders=("order_id", "count"),
        pincode_rto=("rto_flag", "sum"),
    ).reset_index()
    p["pincode_rto_rate"] = (p["pincode_rto"] / p["pincode_orders"]).fillna(0)
    fact = fact.merge(p[["shipping_pincode", "pincode_rto_rate"]], on="shipping_pincode", how="left")
    fact["pincode_rto_rate"] = fact["pincode_rto_rate"].fillna(0)
    fact["high_rto_pincode_flag"] = (fact["pincode_rto_rate"] >= 0.25).astype(int)

    fact["top_category"] = fact["top_category"].astype("string").fillna("unknown")
    fact["net_amount"] = pd.to_numeric(fact["net_amount"], errors="coerce").fillna(0)
    fact["net_amount_z"] = zscore_within_group(fact, "top_category", "net_amount")
    fact["value_outlier_flag"] = (fact["net_amount_z"].abs() >= 2.5).astype(int)

    return fact


# =========================
# Weekly + Queue
# =========================
def build_user_weekly(fact: pd.DataFrame) -> pd.DataFrame:
    df = fact.copy()
    if "order_ts" not in df.columns:
        df["order_ts"] = pd.NaT
    df["week_start"] = week_start(df["order_ts"]).astype("string")

    out = df.groupby(["user_id", "week_start"], as_index=False).agg(
        orders_count=("order_id", "count"),
        net_revenue=("net_amount", "sum"),
        refunds_count=("refund_flag", "sum"),
        refund_amount=("refund_amount", "sum"),
        coupon_orders_count=("coupon_used_flag", "sum"),
        avg_discount_pct=("discount_pct", "mean"),
        cod_orders_count=("cod_flag", "sum"),
        rto_count=("rto_flag", "sum"),
        payment_failures_count=("failed_attempts", "sum"),
        risk_score_avg=("risk_score", "mean"),
    )

    out["avg_discount_pct"] = out["avg_discount_pct"].round(2)
    out["risk_score_avg"] = out["risk_score_avg"].round(1)
    return out

def evidence_for_reason(row: pd.Series, reason: str) -> str:
    if reason == "HIGH_DISCOUNT":
        return f"discount_pct={float(pd.to_numeric(row.get('discount_pct', 0), errors='coerce') or 0):.2f}"

    if reason == "DISCOUNT_SEVERITY":
        return f"discount_pct={float(pd.to_numeric(row.get('discount_pct', 0), errors='coerce') or 0):.2f}"

    if reason == "COUPON_REPEAT_USER":
        return f"coupon_orders_user={int(pd.to_numeric(row.get('coupon_orders_user', 0), errors='coerce') or 0)}"

    if reason == "COUPON_DEVICE_REUSE":
        return f"coupon_users_per_device={int(pd.to_numeric(row.get('coupon_users_per_device', 0), errors='coerce') or 0)}"

    if reason == "PAYMENT_FAIL_SPIKE":
        return f"failed_attempts={int(pd.to_numeric(row.get('failed_attempts', 0), errors='coerce') or 0)}"

    if reason == "NEW_USER":
        return f"days_since_signup={int(pd.to_numeric(row.get('days_since_signup', 9999), errors='coerce') or 9999)}"

    if reason == "NEW_USER_COUPON":
        days = int(pd.to_numeric(row.get("days_since_signup", 9999), errors="coerce") or 9999)
        coupon = str(row.get("coupon_id", "none"))
        return f"days_since_signup={days}; coupon_id={coupon}"

    if reason == "COD_ORDER":
        return f"payment_method={row.get('payment_method', 'unknown')}"

    if reason == "HIGH_RTO_PINCODE":
        rate = float(pd.to_numeric(row.get("pincode_rto_rate", 0), errors="coerce") or 0)
        pin = str(row.get("shipping_pincode", "unknown"))
        return f"shipping_pincode={pin}; pincode_rto_rate={rate:.2f}"

    if reason == "PINCODE_REUSE":
        return f"pincode_orders_count={int(pd.to_numeric(row.get('pincode_orders_count', 0), errors='coerce') or 0)}"

    if reason == "DEVICE_REUSE":
        return f"device_orders_count={int(pd.to_numeric(row.get('device_orders_count', 0), errors='coerce') or 0)}"

    if reason == "VALUE_OUTLIER":
        return f"net_amount_z={float(pd.to_numeric(row.get('net_amount_z', 0), errors='coerce') or 0):.2f}"

    if reason == "REFUND_HISTORY":
        return f"prior_refunds_count={int(pd.to_numeric(row.get('prior_refunds_count', 0), errors='coerce') or 0)}"

    return ""


def build_queue(fact: pd.DataFrame) -> pd.DataFrame:
    df = fact.copy()
    df["recommended_action"] = df.apply(recommend_action, axis=1)

    defaults = {
        "discount_pct": 0,
        "failed_attempts": 0,
        "device_orders_count": 1,
        "coupon_orders_user": 0,
        "coupon_users_per_device": 0,
        "pincode_orders_count": 1,
        "days_since_signup": 9999,
        "prior_refunds_count": 0,
        "shipping_pincode": "unknown",
        "coupon_id": "none",
        "payment_method": "unknown",
        "net_amount": 0,
        "net_amount_z": 0,
        "pincode_rto_rate": 0,
        "order_ts": pd.NaT,
    }
    for c, d in defaults.items():
        if c not in df.columns:
            df[c] = d

    df["evidence_1"] = df.apply(lambda r: evidence_for_reason(r, r.get("reason_1", "")), axis=1)
    df["evidence_2"] = df.apply(lambda r: evidence_for_reason(r, r.get("reason_2", "")), axis=1)
    df["evidence_3"] = df.apply(lambda r: evidence_for_reason(r, r.get("reason_3", "")), axis=1)

    queue = df[[
        "order_id", "user_id", "order_ts", "net_amount", "payment_method", "coupon_id",
        "shipping_pincode",
        "risk_score", "risk_band",
        "reason_1", "evidence_1",
        "reason_2", "evidence_2",
        "reason_3", "evidence_3",
        "recommended_action",
        "discount_pct", "failed_attempts", "device_orders_count", "pincode_orders_count",
        "days_since_signup", "prior_refunds_count",
    ]].copy()

    queue = queue.sort_values(["risk_score", "net_amount"], ascending=[False, False]).reset_index(drop=True)
    queue.insert(0, "rank", range(1, len(queue) + 1))
    return queue

# =========================
# Analysis artifacts
# =========================
def build_analysis_artifacts(fact: pd.DataFrame, weekly: pd.DataFrame, queue: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    # Weekly KPI table (for dashboard + report)
    kpi_weekly = weekly.groupby("week_start", as_index=False).agg(
        orders=("orders_count", "sum"),
        net_revenue=("net_revenue", "sum"),
        refunds=("refunds_count", "sum"),
        refund_amount=("refund_amount", "sum"),
        rto=("rto_count", "sum"),
        coupon_orders=("coupon_orders_count", "sum"),
        payment_failures=("payment_failures_count", "sum"),
        avg_risk=("risk_score_avg", "mean"),
    )
    kpi_weekly["refund_rate"] = (kpi_weekly["refunds"] / kpi_weekly["orders"]).replace([np.inf, np.nan], 0).round(4)
    kpi_weekly["rto_rate"] = (kpi_weekly["rto"] / kpi_weekly["orders"]).replace([np.inf, np.nan], 0).round(4)

    # Top patterns (simple but strong)
    # Pattern 1: coupon_id
    by_coupon = fact.groupby("coupon_id", as_index=False).agg(
        orders=("order_id", "count"),
        refund_amount=("refund_amount", "sum"),
        rto=("rto_flag", "sum"),
        avg_risk=("risk_score", "mean"),
        avg_discount=("discount_pct", "mean"),
    ).sort_values(["avg_risk", "orders"], ascending=[False, False])

    # Pattern 2: pincode
    by_pincode = fact.groupby("shipping_pincode", as_index=False).agg(
        orders=("order_id", "count"),
        refund_amount=("refund_amount", "sum"),
        rto=("rto_flag", "sum"),
        avg_risk=("risk_score", "mean"),
        rto_rate=("pincode_rto_rate", "mean"),
    ).sort_values(["avg_risk", "orders"], ascending=[False, False])

    # Pattern 3: device_id (if present)
    if "device_id" in fact.columns:
        by_device = fact.groupby("device_id", as_index=False).agg(
            orders=("order_id", "count"),
            users=("user_id", "nunique"),
            refund_amount=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["users", "avg_risk"], ascending=[False, False])
    else:
        by_device = pd.DataFrame(columns=["device_id", "orders", "users", "refund_amount", "avg_risk"])

    # Summarize patterns (top 10 each)
    patterns_summary = pd.concat([
        by_coupon.head(10).assign(pattern_type="coupon"),
        by_pincode.head(10).assign(pattern_type="pincode"),
        by_device.head(10).assign(pattern_type="device"),
    ], ignore_index=True)

    return {
        "kpi_weekly": kpi_weekly,
        "patterns_summary": patterns_summary,
        "by_coupon": by_coupon,
        "by_pincode": by_pincode,
        "by_device": by_device,
    }


def export_analysis_report_html(
    kpi_weekly: pd.DataFrame,
    patterns_summary: pd.DataFrame,
    out_path: Path,
    weekly_diagnosis: pd.DataFrame | None = None,
    named_patterns: pd.DataFrame | None = None,
    investigation_table: pd.DataFrame | None = None,
    segment_outputs: Dict[str, pd.DataFrame] | None = None,
) -> None:
    def safe_html(df: pd.DataFrame, max_rows: int | None = None) -> str:
        if df is None or df.empty:
            return "<p><i>No data available.</i></p>"
        x = df.copy()
        if max_rows:
            x = x.head(max_rows)
        return x.to_html(index=False, na_rep="", border=1)

    # Split pattern tables cleanly
    by_coupon = pd.DataFrame()
    by_pincode = pd.DataFrame()
    by_device = pd.DataFrame()

    if patterns_summary is not None and not patterns_summary.empty and "pattern_type" in patterns_summary.columns:
        by_coupon = patterns_summary[patterns_summary["pattern_type"] == "coupon"].copy()
        by_pincode = patterns_summary[patterns_summary["pattern_type"] == "pincode"].copy()
        by_device = patterns_summary[patterns_summary["pattern_type"] == "device"].copy()

    # Executive commentary from current metrics
    summary_points = []
    if kpi_weekly is not None and not kpi_weekly.empty:
        top_refund_week = kpi_weekly.sort_values("refund_amount", ascending=False).head(1)
        if not top_refund_week.empty:
            r = top_refund_week.iloc[0]
            summary_points.append(
                f"Highest weekly refund loss was in <b>{r['week_start']}</b> with "
                f"<b>₹{float(r['refund_amount']):,.2f}</b> across <b>{int(r['refunds'])}</b> refunded orders."
            )

        avg_refund_rate = float(kpi_weekly["refund_rate"].mean()) if "refund_rate" in kpi_weekly.columns else 0
        summary_points.append(
            f"Average weekly refund rate across the period was <b>{avg_refund_rate * 100:.2f}%</b>."
        )

        if "rto" in kpi_weekly.columns and float(kpi_weekly["rto"].sum()) == 0:
            summary_points.append(
                "No RTO events were observed in the current sample, so RTO-based fraud controls cannot yet be validated from this dataset."
            )

        if "payment_failures" in kpi_weekly.columns and float(kpi_weekly["payment_failures"].sum()) == 0:
            summary_points.append(
                "No payment-failure spikes were observed in the weekly rollup, which suggests either low retry abuse in the sample or sparse payment event capture."
            )

    if not by_device.empty:
        top_dev = by_device.sort_values("avg_risk", ascending=False).head(1)
        if not top_dev.empty:
            r = top_dev.iloc[0]
            summary_points.append(
                f"Most suspicious device cluster was <b>{r['device_id']}</b> with "
                f"<b>{int(r['orders'])}</b> orders, <b>{int(r['users'])}</b> users, and average risk score "
                f"<b>{float(r['avg_risk']):.2f}</b>."
            )

    if not by_coupon.empty:
        top_coupon = by_coupon.sort_values("avg_risk", ascending=False).head(1)
        if not top_coupon.empty:
            r = top_coupon.iloc[0]
            summary_points.append(
                f"Highest-risk coupon pattern was <b>{r['coupon_id']}</b> with "
                f"<b>{int(r['orders'])}</b> orders and average risk score <b>{float(r['avg_risk']):.2f}</b>."
            )

    if not by_pincode.empty:
        top_pin = by_pincode.sort_values("refund_amount", ascending=False).head(1)
        if not top_pin.empty:
            r = top_pin.iloc[0]
            summary_points.append(
                f"Highest-loss pincode cluster was <b>{r['shipping_pincode']}</b> with "
                f"loss proxy of <b>₹{float(r['refund_amount']):,.2f}</b> across <b>{int(r['orders'])}</b> orders."
            )

    # Segment deep dive section
    segment_html = ""
    if segment_outputs:
        for name, df in segment_outputs.items():
            pretty_name = name.replace("_", " ").title()
            segment_html += f"<h3>{pretty_name}</h3>\n"
            segment_html += safe_html(df, max_rows=10)

    html = f"""
    <html>
    <head>
        <meta charset="utf-8">
        <title>Fraud Monitoring Analysis Report</title>
        <style>
            body {{
                font-family: Arial, sans-serif;
                margin: 24px;
                color: #222;
            }}
            h1 {{
                color: #1f4e79;
            }}
            h2 {{
                color: #2f5597;
                margin-top: 28px;
            }}
            h3 {{
                color: #444;
                margin-top: 18px;
            }}
            table {{
                border-collapse: collapse;
                width: 100%;
                margin-top: 10px;
                margin-bottom: 20px;
                font-size: 13px;
            }}
            th, td {{
                border: 1px solid #ccc;
                padding: 6px 8px;
                text-align: left;
                vertical-align: top;
            }}
            th {{
                background-color: #f2f2f2;
            }}
            ul {{
                margin-top: 8px;
            }}
            .summary-box {{
                background: #f8fbff;
                border: 1px solid #d6e6f5;
                padding: 14px 16px;
                margin-bottom: 20px;
            }}
            .note-box {{
                background: #fafafa;
                border-left: 4px solid #999;
                padding: 10px 14px;
                margin-top: 16px;
            }}
        </style>
    </head>
    <body>
        <h1>Fraud Monitoring – Analysis Report</h1>

        <div class="summary-box">
            <h2 style="margin-top:0;">Executive Summary</h2>
            <ul>
                {''.join(f'<li>{pt}</li>' for pt in summary_points) if summary_points else '<li>No summary insights available.</li>'}
            </ul>
        </div>

        <h2>1. Weekly KPI Summary</h2>
        {safe_html(kpi_weekly)}

        <h2>2. Weekly Diagnosis / Spike Review</h2>
        {safe_html(weekly_diagnosis)}

        <h2>3. Top Coupon Patterns</h2>
        {safe_html(by_coupon, max_rows=10)}

        <h2>4. Top Pincode Patterns</h2>
        {safe_html(by_pincode, max_rows=10)}

        <h2>5. Top Device Patterns</h2>
        {safe_html(by_device, max_rows=10)}

        <h2>6. Named Fraud / Anomaly Patterns</h2>
        {safe_html(named_patterns, max_rows=10)}

        <h2>7. Segment Deep Dive</h2>
        {segment_html if segment_html else "<p><i>No segment deep-dive outputs available.</i></p>"}

        <h2>8. Investigation Table for Top 2 Patterns</h2>
        {safe_html(investigation_table, max_rows=10)}

        <h2>9. Notes & Limitations</h2>
        <div class="note-box">
            <ul>
                <li>Refund amount is treated as a proxy for fraud loss; it is not a confirmed fraud label.</li>
                <li>RTO and payment-failure trends are currently flat in this sample, so those patterns are weakly represented.</li>
                <li>Risk score is explainable through reason_1, reason_2, and reason_3 derived from weighted signals.</li>
                <li>Some segment views may show <i>unknown</i> if source fields such as channel or city tier are not available in raw data.</li>
            </ul>
        </div>
    </body>
    </html>
    """
    out_path.write_text(html, encoding="utf-8")
    print(f"✅ Wrote: {out_path}")

def label_user_type(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "new_user_flag" in out.columns:
        out["user_type"] = np.where(pd.to_numeric(out["new_user_flag"], errors="coerce").fillna(0) == 1, "new", "returning")
    else:
        out["user_type"] = "unknown"
    return out


def derive_city_tier(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    if "shipping_city_tier" not in out.columns:
        out["shipping_city_tier"] = "unknown"
    out["shipping_city_tier"] = out["shipping_city_tier"].astype("string").fillna("unknown")
    return out


def add_payment_group(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    pm = out.get("payment_method", pd.Series(["unknown"] * len(out), index=out.index)).astype("string").str.lower()
    out["payment_group"] = np.where(pm.str.contains("cod", na=False), "COD", "Prepaid")
    return out


def build_weekly_diagnosis(fact: pd.DataFrame, kpi_weekly: pd.DataFrame) -> pd.DataFrame:
    df = fact.copy()
    wk = kpi_weekly.copy()

    if wk.empty:
        return pd.DataFrame(columns=[
            "week_start", "refund_amount", "rto", "coupon_orders", "payment_failures",
            "spike_metric", "spike_flag", "likely_causes"
        ])

    for col in ["refund_amount", "rto", "coupon_orders", "payment_failures"]:
        if col not in wk.columns:
            wk[col] = 0

    for col in ["refund_amount", "rto", "coupon_orders", "payment_failures"]:
        base = wk[col].rolling(window=3, min_periods=1).mean().shift(1)
        ratio = np.where((base.fillna(0) > 0), wk[col] / base, 1.0)
        wk[f"{col}_vs_prev3_ratio"] = np.round(ratio, 2)

    def detect_spike(row):
        candidates = {
            "refund_amount": row.get("refund_amount_vs_prev3_ratio", 1.0),
            "rto": row.get("rto_vs_prev3_ratio", 1.0),
            "coupon_orders": row.get("coupon_orders_vs_prev3_ratio", 1.0),
            "payment_failures": row.get("payment_failures_vs_prev3_ratio", 1.0),
        }
        metric = max(candidates, key=candidates.get)
        spike = candidates[metric] >= 1.5
        return pd.Series([metric, int(spike)])

    wk[["spike_metric", "spike_flag"]] = wk.apply(detect_spike, axis=1)

    causes = []
    for _, row in wk.iterrows():
        week = row["week_start"]
        sub = df.copy()
        if "order_ts" in sub.columns:
            sub["week_start"] = week_start(sub["order_ts"]).astype("string")
            sub = sub[sub["week_start"] == str(week)]

        reason_cols = [c for c in ["reason_1", "reason_2", "reason_3"] if c in sub.columns]
        if reason_cols and not sub.empty:
            rs = pd.concat([sub[c].astype("string") for c in reason_cols], ignore_index=True)
            rs = rs[rs.notna() & (rs != "")]
            top_reason = rs.value_counts().index[0] if not rs.empty else "NO_CLEAR_REASON"
        else:
            top_reason = "NO_CLEAR_REASON"

        causes.append(top_reason)

    wk["likely_causes"] = causes
    return wk[[
        "week_start", "refund_amount", "rto", "coupon_orders", "payment_failures",
        "spike_metric", "spike_flag", "likely_causes"
    ]]

def build_named_patterns(fact: pd.DataFrame) -> pd.DataFrame:
    df = fact.copy()
    df = label_user_type(df)
    df = add_payment_group(df)
    df = derive_city_tier(df)

    for col in ["refund_amount", "risk_score", "discount_pct", "failed_attempts", "rto_flag"]:
        if col not in df.columns:
            df[col] = 0

    patterns = []

    # Pattern 1
    p1 = df[
        (df.get("coupon_used_flag", 0) == 1) &
        (df.get("new_user_flag", 0) == 1)
    ]
    if not p1.empty:
        g = p1.groupby(["coupon_id", "shipping_pincode"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["orders", "loss_proxy"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "coupon used by many new users from same pincode",
                "pattern_definition": f"coupon_id={r['coupon_id']} with new-user concentration in pincode={r['shipping_pincode']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "new users + pincode cluster",
                "segment_value": str(r["shipping_pincode"]),
                "avg_risk": float(r["avg_risk"]),
            })

    # Pattern 2
    p2 = df[df.get("failed_attempts", 0) >= 2]
    if not p2.empty:
        g = p2.groupby(["payment_method"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["orders", "avg_risk"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "many payment failures before order completion",
                "pattern_definition": f"orders with failed_attempts>=2 concentrated in payment_method={r['payment_method']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "payment method",
                "segment_value": str(r["payment_method"]),
                "avg_risk": float(r["avg_risk"]),
            })

    # Pattern 3
    p3 = df[(df.get("payment_group", "") == "COD") & (df.get("rto_flag", 0) == 1)]
    if not p3.empty:
        g = p3.groupby(["shipping_pincode"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["orders", "avg_risk"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "COD orders from specific region with high RTO",
                "pattern_definition": f"COD RTO concentration in pincode={r['shipping_pincode']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "pincode",
                "segment_value": str(r["shipping_pincode"]),
                "avg_risk": float(r["avg_risk"]),
            })

    # Pattern 4
    if "device_id" in df.columns:
        p4 = df[df.get("device_orders_count", 0) >= 3]
        if not p4.empty:
            g = p4.groupby(["device_id"], as_index=False).agg(
                orders=("order_id", "count"),
                users=("user_id", "nunique"),
                loss_proxy=("refund_amount", "sum"),
                avg_risk=("risk_score", "mean"),
            ).sort_values(["users", "avg_risk"], ascending=[False, False]).head(1)
            if not g.empty:
                r = g.iloc[0]
                patterns.append({
                    "pattern_name": "same device reused across multiple suspicious orders",
                    "pattern_definition": f"device_id={r['device_id']} reused across {int(r['users'])} users",
                    "orders": int(r["orders"]),
                    "loss_proxy": float(r["loss_proxy"]),
                    "segment": "device",
                    "segment_value": str(r["device_id"]),
                    "avg_risk": float(r["avg_risk"]),
                })

    # Pattern 5
    p5 = df[df.get("high_discount_flag", 0) == 1]
    if not p5.empty:
        g = p5.groupby(["top_category"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_discount=("discount_pct", "mean"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["loss_proxy", "orders"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "high-discount concentration in category",
                "pattern_definition": f"discount-heavy suspicious orders concentrated in category={r['top_category']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "category",
                "segment_value": str(r["top_category"]),
                "avg_risk": float(r["avg_risk"]),
            })

    return pd.DataFrame(patterns)

def build_named_patterns(fact: pd.DataFrame) -> pd.DataFrame:
    df = fact.copy()
    df = label_user_type(df)
    df = add_payment_group(df)
    df = derive_city_tier(df)

    for col in ["refund_amount", "risk_score", "discount_pct", "failed_attempts", "rto_flag"]:
        if col not in df.columns:
            df[col] = 0

    patterns = []

    # Pattern 1
    p1 = df[
        (df.get("coupon_used_flag", 0) == 1) &
        (df.get("new_user_flag", 0) == 1)
    ]
    if not p1.empty:
        g = p1.groupby(["coupon_id", "shipping_pincode"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["orders", "loss_proxy"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "coupon used by many new users from same pincode",
                "pattern_definition": f"coupon_id={r['coupon_id']} with new-user concentration in pincode={r['shipping_pincode']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "new users + pincode cluster",
                "segment_value": str(r["shipping_pincode"]),
                "avg_risk": float(r["avg_risk"]),
            })

    # Pattern 2
    p2 = df[df.get("failed_attempts", 0) >= 2]
    if not p2.empty:
        g = p2.groupby(["payment_method"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["orders", "avg_risk"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "many payment failures before order completion",
                "pattern_definition": f"orders with failed_attempts>=2 concentrated in payment_method={r['payment_method']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "payment method",
                "segment_value": str(r["payment_method"]),
                "avg_risk": float(r["avg_risk"]),
            })

    # Pattern 3
    p3 = df[(df.get("payment_group", "") == "COD") & (df.get("rto_flag", 0) == 1)]
    if not p3.empty:
        g = p3.groupby(["shipping_pincode"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["orders", "avg_risk"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "COD orders from specific region with high RTO",
                "pattern_definition": f"COD RTO concentration in pincode={r['shipping_pincode']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "pincode",
                "segment_value": str(r["shipping_pincode"]),
                "avg_risk": float(r["avg_risk"]),
            })

    # Pattern 4
    if "device_id" in df.columns:
        p4 = df[df.get("device_orders_count", 0) >= 3]
        if not p4.empty:
            g = p4.groupby(["device_id"], as_index=False).agg(
                orders=("order_id", "count"),
                users=("user_id", "nunique"),
                loss_proxy=("refund_amount", "sum"),
                avg_risk=("risk_score", "mean"),
            ).sort_values(["users", "avg_risk"], ascending=[False, False]).head(1)
            if not g.empty:
                r = g.iloc[0]
                patterns.append({
                    "pattern_name": "same device reused across multiple suspicious orders",
                    "pattern_definition": f"device_id={r['device_id']} reused across {int(r['users'])} users",
                    "orders": int(r["orders"]),
                    "loss_proxy": float(r["loss_proxy"]),
                    "segment": "device",
                    "segment_value": str(r["device_id"]),
                    "avg_risk": float(r["avg_risk"]),
                })

    # Pattern 5
    p5 = df[df.get("high_discount_flag", 0) == 1]
    if not p5.empty:
        g = p5.groupby(["top_category"], as_index=False).agg(
            orders=("order_id", "count"),
            loss_proxy=("refund_amount", "sum"),
            avg_discount=("discount_pct", "mean"),
            avg_risk=("risk_score", "mean"),
        ).sort_values(["loss_proxy", "orders"], ascending=[False, False]).head(1)
        if not g.empty:
            r = g.iloc[0]
            patterns.append({
                "pattern_name": "high-discount concentration in category",
                "pattern_definition": f"discount-heavy suspicious orders concentrated in category={r['top_category']}",
                "orders": int(r["orders"]),
                "loss_proxy": float(r["loss_proxy"]),
                "segment": "category",
                "segment_value": str(r["top_category"]),
                "avg_risk": float(r["avg_risk"]),
            })

    return pd.DataFrame(patterns)

def top_n_segment_stats(df: pd.DataFrame, group_cols: List[str], n: int = 3) -> tuple[pd.DataFrame, pd.DataFrame]:
    work = df.copy()
    for c in group_cols:
        if c not in work.columns:
            work[c] = "unknown"

    agg = work.groupby(group_cols, as_index=False).agg(
        orders=("order_id", "count"),
        refund_loss=("refund_amount", "sum"),
        avg_risk=("risk_score", "mean"),
        rto=("rto_flag", "sum"),
        coupon_orders=("coupon_used_flag", "sum"),
    )

    top_loss = agg.sort_values(["refund_loss", "orders"], ascending=[False, False]).head(n).copy()
    top_risk = agg.sort_values(["avg_risk", "orders"], ascending=[False, False]).head(n).copy()
    return top_loss, top_risk


def recommend_control_for_segment(row: pd.Series, segment_name: str) -> str:
    seg = str(row.to_dict())
    if "COD" in seg or "rto" in segment_name.lower():
        return "Add COD address confirmation / OTP before shipment."
    if "new" in seg and "coupon" in seg:
        return "Apply coupon throttling and verification for new users."
    if "device" in segment_name.lower():
        return "Throttle repeated device usage and require step-up verification."
    if "payment" in segment_name.lower():
        return "Add payment retry controls / call verification."
    return "Route segment to targeted monitoring and manual review."


def build_segment_deep_dive(fact: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    df = fact.copy()
    df = label_user_type(df)
    df = add_payment_group(df)
    df = derive_city_tier(df)

    if "channel" not in df.columns:
        df["channel"] = "unknown"
    if "device_id" not in df.columns:
        df["device_id"] = "unknown"

    outputs = {}

    analyses = {
        "channel_x_device": ["channel", "device_id"],
        "new_vs_returning": ["user_type"],
        "city_tier_clusters": ["shipping_city_tier", "shipping_pincode"],
        "payment_method_group": ["payment_group"],
        "top_categories": ["top_category"],
    }

    for name, cols in analyses.items():
        top_loss, top_risk = top_n_segment_stats(df, cols, n=3)
        if not top_loss.empty:
            top_loss["recommended_control"] = top_loss.apply(lambda r: recommend_control_for_segment(r, name), axis=1)
        if not top_risk.empty:
            top_risk["recommended_control"] = top_risk.apply(lambda r: recommend_control_for_segment(r, name), axis=1)

        outputs[f"{name}_top_loss"] = top_loss
        outputs[f"{name}_top_risk"] = top_risk

    return outputs

def build_investigation_table(fact: pd.DataFrame, named_patterns: pd.DataFrame) -> pd.DataFrame:
    if named_patterns.empty:
        return pd.DataFrame(columns=[
            "pattern_name", "pattern_definition", "volume_orders", "loss_proxy",
            "concentration", "hypothesis_1", "hypothesis_2", "hypothesis_3",
            "validation_needed", "evidence_used"
        ])

    top2 = named_patterns.sort_values(["loss_proxy", "orders"], ascending=[False, False]).head(2).copy()

    rows = []
    for _, r in top2.iterrows():
        pattern = str(r["pattern_name"])
        definition = str(r["pattern_definition"])
        orders = int(r["orders"])
        loss = float(r["loss_proxy"])
        concentration = f"{r['segment']}={r['segment_value']}"

        if "coupon" in pattern.lower():
            h1 = "Promo code leaked to abuse networks or public forums."
            h2 = "New-user onboarding loophole enables repeated incentive capture."
            h3 = "Weak device / pincode throttling allows repeated redemption."
            validation = "Check coupon redemption history by device, pincode, and signup cohort; compare before/after restriction."
            evidence = "patterns_summary, top_coupons_avg_risk chart, investigation_queue"
        elif "payment failures" in pattern.lower():
            h1 = "Card testing or bot-driven retry behavior before a successful transaction."
            h2 = "Fraudsters probe authorization rules with low-value attempts."
            h3 = "Gateway retry logic may allow suspicious retries without escalation."
            validation = "Inspect payment event sequence and session-level retry patterns before successful orders."
            evidence = "kpi_weekly payment_failures, investigation_queue failed_attempts"
        else:
            h1 = "Regional or operational weaknesses create a repeatable fraud opportunity."
            h2 = "Certain user/device cohorts are exploiting known control gaps."
            h3 = "Model thresholds may not yet block this suspicious segment early enough."
            validation = "Drill into the segment over 2–4 weeks and compare treated vs untreated cohorts."
            evidence = "weekly diagnosis, named patterns, segment deep dive tables"

        rows.append({
            "pattern_name": pattern,
            "pattern_definition": definition,
            "volume_orders": orders,
            "loss_proxy": round(loss, 2),
            "concentration": concentration,
            "hypothesis_1": h1,
            "hypothesis_2": h2,
            "hypothesis_3": h3,
            "validation_needed": validation,
            "evidence_used": evidence,
        })

    return pd.DataFrame(rows)


def build_part_c_outputs(fact: pd.DataFrame, kpi_weekly: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    weekly_diagnosis = build_weekly_diagnosis(fact, kpi_weekly)
    named_patterns = build_named_patterns(fact)
    segment_outputs = build_segment_deep_dive(fact)
    investigation_table = build_investigation_table(fact, named_patterns)

    outputs = {
        "weekly_diagnosis": weekly_diagnosis,
        "named_patterns": named_patterns,
        "investigation_table_top2_patterns": investigation_table,
    }
    outputs.update(segment_outputs)
    return outputs
# =========================
# Dashboard generation (Excel + images)
# =========================
def export_charts(kpi_weekly: pd.DataFrame, fact: pd.DataFrame) -> None:
    # 1) Refund amount trend
    if not kpi_weekly.empty and "week_start" in kpi_weekly.columns:
        plt.figure(figsize=(8, 5))
        plt.plot(kpi_weekly["week_start"], kpi_weekly["refund_amount"])
        plt.xticks(rotation=45, ha="right")
        plt.title("Weekly Refund Amount Trend")
        plt.tight_layout()
        path1 = EXPORTS_DIR / "weekly_refunds_trend.png"
        plt.savefig(path1, dpi=160)
        plt.close()
        print(f"✅ Wrote: {path1}")

    # 2) Risk band distribution
    if "risk_band" in fact.columns:
        band_order = ["Low", "Medium", "High"]
        band_counts = (
            fact["risk_band"]
            .astype("string")
            .value_counts(dropna=False)
            .reindex(band_order, fill_value=0)
        )

        plt.figure(figsize=(8, 5))
        bars = plt.bar(band_counts.index, band_counts.values)

        for bar, val in zip(bars, band_counts.values):
            plt.text(
                bar.get_x() + bar.get_width() / 2,
                bar.get_height(),
                str(int(val)),
                ha="center",
                va="bottom"
            )

        plt.title("Risk Band Distribution (All Orders)")
        plt.ylabel("Order Count")
        plt.tight_layout()
        path2 = EXPORTS_DIR / "risk_band_split.png"
        plt.savefig(path2, dpi=160)
        plt.close()
        print(f"✅ Wrote: {path2}")

    # 3) Top coupons by avg risk
    if "coupon_id" in fact.columns and "risk_score" in fact.columns:
        tmp = (
            fact.groupby("coupon_id", as_index=False)
            .agg(avg_risk=("risk_score", "mean"), orders=("order_id", "count"))
            .sort_values(["avg_risk", "orders"], ascending=[False, False])
            .head(10)
        )
        if not tmp.empty:
            plt.figure(figsize=(8, 5))
            plt.bar(tmp["coupon_id"].astype(str), tmp["avg_risk"])
            plt.xticks(rotation=45, ha="right")
            plt.title("Top Coupons by Average Risk")
            plt.tight_layout()
            path3 = EXPORTS_DIR / "top_coupons_avg_risk.png"
            plt.savefig(path3, dpi=160)
            plt.close()
            print(f"✅ Wrote: {path3}")

    # 4) Payment method avg risk
    if "payment_method" in fact.columns and "risk_score" in fact.columns:
        tmp = (
            fact.groupby("payment_method", as_index=False)
            .agg(avg_risk=("risk_score", "mean"), orders=("order_id", "count"))
            .sort_values(["avg_risk", "orders"], ascending=[False, False])
        )
        if not tmp.empty:
            plt.figure(figsize=(8, 5))
            plt.bar(tmp["payment_method"].astype(str), tmp["avg_risk"])
            plt.xticks(rotation=45, ha="right")
            plt.title("Average Risk by Payment Method")
            plt.tight_layout()
            path4 = EXPORTS_DIR / "avg_risk_by_payment_method.png"
            plt.savefig(path4, dpi=160)
            plt.close()
            print(f"✅ Wrote: {path4}")

def autosize_excel_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(45, max(10, max_len + 2))

def build_dashboard_views(fact: pd.DataFrame, queue: pd.DataFrame, kpi_weekly: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    df = fact.copy()
    q = queue.copy()

    # -------- Executive Summary --------
    total_orders = len(df)
    total_refund_loss = pd.to_numeric(df.get("refund_amount", 0), errors="coerce").fillna(0).sum()
    total_rto = pd.to_numeric(df.get("rto_flag", 0), errors="coerce").fillna(0).sum()
    total_payment_fails = pd.to_numeric(df.get("failed_attempts", 0), errors="coerce").fillna(0).sum()
    coupon_orders = pd.to_numeric(df.get("coupon_used_flag", 0), errors="coerce").fillna(0).sum()

    flagged_orders = int((df.get("risk_score", 0) >= 35).sum()) if "risk_score" in df.columns else 0
    high_risk_orders = int((df.get("risk_band", "") == "High").sum()) if "risk_band" in df.columns else 0
    medium_risk_orders = int((df.get("risk_band", "") == "Medium").sum()) if "risk_band" in df.columns else 0
    low_risk_orders = int((df.get("risk_band", "") == "Low").sum()) if "risk_band" in df.columns else 0

    payment_fail_rate = (total_payment_fails / total_orders) if total_orders else 0
    coupon_abuse_rate = (coupon_orders / total_orders) if total_orders else 0
    rto_rate = (total_rto / total_orders) if total_orders else 0

    executive_summary = pd.DataFrame([
        ["Total Orders", total_orders],
        ["Refund Loss", round(float(total_refund_loss), 2)],
        ["RTO Rate", round(float(rto_rate), 4)],
        ["Payment Fail Rate", round(float(payment_fail_rate), 4)],
        ["Coupon Abuse Rate", round(float(coupon_abuse_rate), 4)],
        ["Flagged Orders", flagged_orders],
        ["High Risk Orders", high_risk_orders],
        ["Medium Risk Orders", medium_risk_orders],
        ["Low Risk Orders", low_risk_orders],
    ], columns=["metric", "value"])

    # -------- Fraud Drivers --------
    by_payment_method = df.groupby("payment_method", as_index=False).agg(
        orders=("order_id", "count"),
        avg_risk=("risk_score", "mean"),
        refund_amount=("refund_amount", "sum"),
        rto=("rto_flag", "sum"),
    ).sort_values(["avg_risk", "orders"], ascending=[False, False]) if "payment_method" in df.columns else pd.DataFrame()

    by_channel = df.groupby("channel", as_index=False).agg(
        orders=("order_id", "count"),
        avg_risk=("risk_score", "mean"),
        refund_amount=("refund_amount", "sum"),
    ).sort_values(["avg_risk", "orders"], ascending=[False, False]) if "channel" in df.columns else pd.DataFrame(
        columns=["channel", "orders", "avg_risk", "refund_amount"]
    )

    by_device = df.groupby("device_id", as_index=False).agg(
        orders=("order_id", "count"),
        users=("user_id", "nunique"),
        avg_risk=("risk_score", "mean"),
        refund_amount=("refund_amount", "sum"),
    ).sort_values(["avg_risk", "orders"], ascending=[False, False]) if "device_id" in df.columns else pd.DataFrame(
        columns=["device_id", "orders", "users", "avg_risk", "refund_amount"]
    )

    by_coupon = df.groupby("coupon_id", as_index=False).agg(
        orders=("order_id", "count"),
        avg_risk=("risk_score", "mean"),
        refund_amount=("refund_amount", "sum"),
    ).sort_values(["avg_risk", "orders"], ascending=[False, False]) if "coupon_id" in df.columns else pd.DataFrame(
        columns=["coupon_id", "orders", "avg_risk", "refund_amount"]
    )

    # -------- Operational / Queue --------
    queue_summary = q.groupby(["risk_band", "recommended_action"], as_index=False).agg(
        orders=("order_id", "count"),
        total_value=("net_amount", "sum"),
    ).sort_values(["risk_band", "orders"], ascending=[True, False]) if not q.empty else pd.DataFrame()

    # -------- Controls & Impact --------
    manual_review_volume = int(q["recommended_action"].isin(["MANUAL_REVIEW", "HOLD_FOR_MANUAL_REVIEW"]).sum()) if not q.empty else 0
    otp_volume = int(q["recommended_action"].isin(["SOFT_FRICTION_OTP", "OTP_ADDRESS_CONFIRMATION"]).sum()) if not q.empty else 0
    call_verify_volume = int((q["recommended_action"] == "CALL_VERIFICATION").sum()) if not q.empty else 0

    refund_loss = float(total_refund_loss)
    preventable_share = 0.20
    capture_rate = 0.50
    projected_prevented_loss = refund_loss * preventable_share * capture_rate

    controls_impact = pd.DataFrame([
        ["Projected Prevented Loss", round(projected_prevented_loss, 2)],
        ["Expected Manual Review Volume", manual_review_volume],
        ["Expected OTP Volume", otp_volume],
        ["Expected Call Verification Volume", call_verify_volume],
        ["Experiment Window (weeks)", "2-4"],
        ["Primary KPI 1", "Refund Rate"],
        ["Primary KPI 2", "RTO Rate"],
        ["Primary KPI 3", "Flagged Order Share"],
        ["Guardrail 1", "Conversion Rate"],
        ["Guardrail 2", "Payment Success Rate"],
    ], columns=["metric", "value"])

    experiment_plan = pd.DataFrame([
        ["Coupon restriction by device/pincode", "High coupon/device reuse", "Refund rate, coupon abuse rate"],
        ["OTP for new-user coupon orders", "New user + coupon risk", "Fraud reduction vs conversion impact"],
        ["COD address confirmation", "High RTO pincode + COD", "RTO rate, delivery success"],
        ["Manual review for High risk", "High risk orders", "Prevented loss, analyst hit rate"],
    ], columns=["control", "target_segment", "success_metrics"])

    return {
        "executive_summary": executive_summary,
        "by_payment_method": by_payment_method,
        "by_channel": by_channel,
        "by_device": by_device,
        "by_coupon_dashboard": by_coupon,
        "queue_summary": queue_summary,
        "controls_impact": controls_impact,
        "experiment_plan": experiment_plan,
    }

def export_dashboard_xlsx(
    fact: pd.DataFrame,
    weekly: pd.DataFrame,
    queue: pd.DataFrame,
    kpi_weekly: pd.DataFrame,
    patterns_summary: pd.DataFrame
) -> None:
    xlsx_path = DASHBOARD_DIR / "dashboard.xlsx"

    dashboard_views = build_dashboard_views(fact, queue, kpi_weekly)

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        # View 1
        dashboard_views["executive_summary"].to_excel(writer, sheet_name="Executive_Summary", index=False)
        kpi_weekly.to_excel(writer, sheet_name="KPI_Weekly", index=False)

        # View 2
        patterns_summary.to_excel(writer, sheet_name="Patterns", index=False)
        dashboard_views["by_coupon_dashboard"].head(50).to_excel(writer, sheet_name="Top_Coupons", index=False)
        dashboard_views["by_payment_method"].head(50).to_excel(writer, sheet_name="By_Payment_Method", index=False)
        dashboard_views["by_device"].head(50).to_excel(writer, sheet_name="By_Device", index=False)
        dashboard_views["by_channel"].head(50).to_excel(writer, sheet_name="By_Channel", index=False)

        # View 3
        queue.head(1000).to_excel(writer, sheet_name="Investigation_Queue", index=False)
        dashboard_views["queue_summary"].to_excel(writer, sheet_name="Queue_Summary", index=False)

        # View 4
        dashboard_views["controls_impact"].to_excel(writer, sheet_name="Controls_Impact", index=False)
        dashboard_views["experiment_plan"].to_excel(writer, sheet_name="Experiment_Plan", index=False)

        # Optional debug sheet
        fact.head(2000).to_excel(writer, sheet_name="Sample_Fact", index=False)

    wb = load_workbook(xlsx_path)

    for sheet in wb.sheetnames:
        ws = wb[sheet]

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        autosize_excel_columns(ws)

    wb.save(xlsx_path)
    print(f"✅ Wrote: {xlsx_path}")

# =========================
# Final story: PDF memo
# =========================
def export_final_memo_pdf(
    fact: pd.DataFrame,
    kpi_weekly: pd.DataFrame,
    patterns_summary: pd.DataFrame,
    queue: pd.DataFrame,
    out_path: Path
) -> None:
    from pathlib import Path
    from typing import List
    import pandas as pd
    import numpy as np
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import inch

    # -----------------------------
    # Config copied here to keep PDF self-contained
    # -----------------------------
    WEIGHTS = {
        "high_discount_flag": 12,
        "discount_severity_score": 10,
        "coupon_repeat_user_flag": 10,
        "coupon_device_reuse_flag": 12,
        "payment_failed_attempts_score": 18,
        "new_user_flag": 8,
        "new_user_plus_coupon": 10,
        "cod_flag": 6,
        "high_rto_pincode_flag": 12,
        "pincode_reuse_score": 10,
        "device_reuse_score": 12,
        "value_outlier_flag": 8,
        "refund_history_user_flag": 10,
    }

    RISK_BAND_RULE = [
        "Low: 0–34",
        "Medium: 35–59",
        "High: 60–100",
    ]

    # -----------------------------
    # Helpers
    # -----------------------------
    def fmt_num(x, decimals=0):
        try:
            if pd.isna(x):
                return "0"
            if decimals == 0:
                return f"{float(x):,.0f}"
            return f"{float(x):,.{decimals}f}"
        except Exception:
            return str(x)

    def fmt_pct(x, decimals=2):
        try:
            if pd.isna(x):
                return "0.00%"
            return f"{float(x) * 100:.{decimals}f}%"
        except Exception:
            return "0.00%"

    def wrap_text(c: canvas.Canvas, text: str, font_name: str, font_size: int, max_width: float) -> List[str]:
        c.setFont(font_name, font_size)
        words = (text or "").split()
        if not words:
            return [""]

        lines = []
        cur = words[0]
        for w in words[1:]:
            trial = cur + " " + w
            if c.stringWidth(trial, font_name, font_size) <= max_width:
                cur = trial
            else:
                lines.append(cur)
                cur = w
        lines.append(cur)
        return lines

    def ensure_space(c: canvas.Canvas, y: float, needed: float):
        bottom_margin = 0.65 * inch
        if y - needed < bottom_margin:
            c.showPage()
            return page_h - top_margin
        return y

    def draw_paragraph(
        c: canvas.Canvas,
        x: float,
        y: float,
        text: str,
        font_name: str = "Helvetica",
        font_size: int = 10,
        line_gap: int = 3
    ) -> float:
        max_width = page_w - left_margin - right_margin
        lines = wrap_text(c, text, font_name, font_size, max_width)
        line_height = font_size + line_gap

        y = ensure_space(c, y, line_height * len(lines))
        c.setFont(font_name, font_size)
        for line in lines:
            c.drawString(x, y, line)
            y -= line_height
        return y

    def draw_heading(c: canvas.Canvas, x: float, y: float, text: str, level: int = 1) -> float:
        if level == 1:
            return draw_paragraph(c, x, y, text, font_name="Helvetica-Bold", font_size=15, line_gap=5)
        if level == 2:
            return draw_paragraph(c, x, y, text, font_name="Helvetica-Bold", font_size=12, line_gap=4)
        return draw_paragraph(c, x, y, text, font_name="Helvetica-Bold", font_size=10, line_gap=3)

    def draw_bullets(
        c: canvas.Canvas,
        x: float,
        y: float,
        bullets: List[str],
        font_name: str = "Helvetica",
        font_size: int = 10,
        bullet_indent: float = 12,
        line_gap: int = 3
    ) -> float:
        max_width = page_w - left_margin - right_margin - bullet_indent
        line_height = font_size + line_gap
        c.setFont(font_name, font_size)

        for b in bullets:
            lines = wrap_text(c, b, font_name, font_size, max_width)
            needed = line_height * max(1, len(lines)) + 4
            y = ensure_space(c, y, needed)

            c.drawString(x, y, "•")
            c.drawString(x + bullet_indent, y, lines[0])
            y -= line_height

            for line in lines[1:]:
                c.drawString(x + bullet_indent, y, line)
                y -= line_height

        return y

    def draw_kv_list(c: canvas.Canvas, x: float, y: float, rows: List[tuple], font_size: int = 10) -> float:
        for k, v in rows:
            y = draw_paragraph(c, x, y, f"{k}: {v}", font_name="Helvetica", font_size=font_size, line_gap=3)
        return y

    # -----------------------------
    # Derived metrics
    # -----------------------------
    total_orders = len(fact)
    total_refund = float(pd.to_numeric(fact.get("refund_amount", 0), errors="coerce").fillna(0).sum())
    total_refunds_count = int(pd.to_numeric(fact.get("refund_flag", 0), errors="coerce").fillna(0).sum())
    total_rto = int(pd.to_numeric(fact.get("rto_flag", 0), errors="coerce").fillna(0).sum())

    high_risk = int((fact.get("risk_band", pd.Series(dtype="string")) == "High").sum()) if "risk_band" in fact.columns else 0
    medium_risk = int((fact.get("risk_band", pd.Series(dtype="string")) == "Medium").sum()) if "risk_band" in fact.columns else 0
    low_risk = int((fact.get("risk_band", pd.Series(dtype="string")) == "Low").sum()) if "risk_band" in fact.columns else 0

    refund_rate = (total_refunds_count / total_orders) if total_orders else 0
    rto_rate = (total_rto / total_orders) if total_orders else 0
    high_risk_rate = (high_risk / total_orders) if total_orders else 0

    avg_risk = float(pd.to_numeric(fact.get("risk_score", 0), errors="coerce").fillna(0).mean()) if total_orders else 0

    top_queue = queue.head(5).copy() if not queue.empty else pd.DataFrame()

    # top patterns
    top_patterns = patterns_summary.head(8).copy() if not patterns_summary.empty else pd.DataFrame()

    # top reasons
    reason_cols = [c for c in ["reason_1", "reason_2", "reason_3"] if c in fact.columns]
    if reason_cols:
        reason_series = pd.concat([fact[c].astype("string") for c in reason_cols], ignore_index=True)
        reason_series = reason_series[reason_series.notna() & (reason_series != "")]
        top_reasons = reason_series.value_counts().head(6)
    else:
        top_reasons = pd.Series(dtype="int64")

    # top suspicious coupon / pincode / device
    def pick_top(df: pd.DataFrame, ptype: str):
        if df.empty or "pattern_type" not in df.columns:
            return None
        x = df[df["pattern_type"] == ptype]
        if x.empty:
            return None
        return x.iloc[0].to_dict()

    top_coupon = pick_top(top_patterns, "coupon")
    top_pincode = pick_top(top_patterns, "pincode")
    top_device = pick_top(top_patterns, "device")

    # curated tables summary
    curated_tables = [
        ("fact_orders_enriched", "1 row per order; master scored fact table used for monitoring, ranking, and reporting"),
        ("fact_user_risk_weekly", "user-week aggregation with order count, revenue, refund, coupon, RTO, and average risk metrics"),
        ("investigation_queue", "priority-ranked order queue with risk score, reasons, and recommended action"),
        ("kpi_weekly", "weekly management KPIs such as orders, refund amount, refund rate, RTO rate, and average risk"),
        ("patterns_summary", "top coupon / pincode / device anomaly patterns for leadership review"),
    ]

    # 30-day estimate assumptions
    # Conservative planning assumptions, clearly labeled
    monthly_loss_proxy = total_refund
    preventable_share = 0.20
    expected_capture_rate = 0.50
    estimated_30_day_savings = monthly_loss_proxy * preventable_share * expected_capture_rate

    # -----------------------------
    # PDF setup
    # -----------------------------
    c = canvas.Canvas(str(out_path), pagesize=letter)
    page_w, page_h = letter
    left_margin = 0.75 * inch
    right_margin = 0.75 * inch
    top_margin = 0.75 * inch

    x = left_margin
    y = page_h - top_margin

    # -----------------------------
    # Title
    # -----------------------------
    y = draw_heading(c, x, y, "Fraud / Anomaly Monitoring & Investigation Dashboard", level=1)
    y -= 4
    y = draw_paragraph(
        c, x, y,
        "Final memo generated from ETL outputs. This memo summarizes monitoring objectives, data pipeline, scoring logic, "
        "investigation prioritization, recommended controls, and a 30-day impact estimate."
    )
    y -= 8

    # -----------------------------
    # 1. Objective + KPI definitions
    # -----------------------------
    y = draw_heading(c, x, y, "1. Objective + KPI Definitions", level=2)
    y = draw_paragraph(
        c, x, y,
        "Objective: build an always-on fraud / anomaly monitoring system that identifies the biggest abnormal patterns, "
        "prioritizes orders for investigation, estimates loss exposure, and supports measurable control experiments."
    )
    y -= 2

    kpi_rows = [
        ("Total Orders", fmt_num(total_orders)),
        ("Refund Amount (proxy loss)", fmt_num(total_refund, 2)),
        ("Refund Rate", fmt_pct(refund_rate)),
        ("RTO Count", fmt_num(total_rto)),
        ("RTO Rate", fmt_pct(rto_rate)),
        ("Average Risk Score", fmt_num(avg_risk, 1)),
        ("High-Risk Orders", f"{fmt_num(high_risk)} ({fmt_pct(high_risk_rate)})"),
    ]
    y = draw_kv_list(c, x, y, kpi_rows)
    y -= 8

    # -----------------------------
    # 2. Data pipeline summary + curated tables
    # -----------------------------
    y = draw_heading(c, x, y, "2. Data Pipeline Summary + Curated Tables", level=2)
    pipeline_bullets = [
        "Raw inputs loaded from users, sessions, orders, order_items, payments, shipments, refunds, coupons, and products.",
        "Data standardized into snake_case, trimmed, parsed for timestamps, and deduplicated on business keys.",
        "A scored fact table is created at order level by joining user, session, payment, shipment, coupon, refund, and product-derived signals.",
        "Weekly KPI tables, pattern summaries, and an investigation queue are generated for dashboarding and operations.",
    ]
    y = draw_bullets(c, x, y, pipeline_bullets)

    y -= 4
    y = draw_paragraph(c, x, y, "Curated output tables:", font_name="Helvetica-Bold", font_size=10)
    table_bullets = [f"{name} — {desc}" for name, desc in curated_tables]
    y = draw_bullets(c, x, y, table_bullets)
    y -= 8

    # -----------------------------
    # 3. Top insights & patterns (quantified)
    # -----------------------------
    y = draw_heading(c, x, y, "3. Top Insights & Patterns (Quantified)", level=2)

    insight_bullets = [
        f"Total proxy loss from refunds is {fmt_num(total_refund, 2)} across {fmt_num(total_refunds_count)} refunded orders.",
        f"Average order risk score is {fmt_num(avg_risk, 1)}; risk-band mix is Low={fmt_num(low_risk)}, Medium={fmt_num(medium_risk)}, High={fmt_num(high_risk)}.",
    ]

    if top_coupon is not None:
        insight_bullets.append(
            f"Top coupon pattern: {top_coupon.get('coupon_id', 'NA')} with {fmt_num(top_coupon.get('orders', 0))} orders, "
            f"avg risk {fmt_num(top_coupon.get('avg_risk', 0), 1)}, and refund amount {fmt_num(top_coupon.get('refund_amount', 0), 2)}."
        )
    if top_pincode is not None:
        insight_bullets.append(
            f"Top pincode pattern: {top_pincode.get('shipping_pincode', 'NA')} with {fmt_num(top_pincode.get('orders', 0))} orders, "
            f"avg risk {fmt_num(top_pincode.get('avg_risk', 0), 1)}, and RTO rate {fmt_pct(top_pincode.get('rto_rate', 0))}."
        )
    if top_device is not None:
        insight_bullets.append(
            f"Top device pattern: {top_device.get('device_id', 'NA')} with {fmt_num(top_device.get('orders', 0))} orders across "
            f"{fmt_num(top_device.get('users', 0))} users and avg risk {fmt_num(top_device.get('avg_risk', 0), 1)}."
        )

    if len(top_reasons) > 0:
        top_reason_str = ", ".join([f"{idx} ({val})" for idx, val in top_reasons.items()])
        insight_bullets.append(f"Most common risk reasons across flagged orders are: {top_reason_str}.")

    y = draw_bullets(c, x, y, insight_bullets)
    y -= 8

    # -----------------------------
    # 4. Scoring system
    # -----------------------------
    y = draw_heading(c, x, y, "4. Scoring System (Signals + Thresholds)", level=2)
    y = draw_paragraph(
        c, x, y,
        "Each order receives an explainable risk score from 0 to 100. The score is the weighted sum of binary or scaled signals, "
        "and the top three contributing reasons are stored for analyst review."
    )
    y -= 3

    scoring_bullets = [
        "High discount flag: discount_pct >= 40",
        "Discount severity score: discount_pct scaled between 0 and 80%",
        "Coupon repeat user flag: user has >= 2 coupon-linked orders",
        "Coupon-device reuse flag: same device used by >= 2 users with coupons",
        "Payment failed attempts score: 0 for none, 0.35 for 1 fail, 0.65 for 2 fails, 1.0 for 3+ fails",
        "New user flag: account age <= 14 days at order time",
        "New user + coupon flag: new user and coupon used in same order",
        "COD flag: payment method contains COD",
        "High-RTO pincode flag: pincode has >= 3 orders and RTO rate >= 20%",
        "Pincode reuse score: repeated order density by pincode scaled 0 to 1",
        "Device reuse score: repeated order density by device scaled 0 to 1",
        "Value outlier flag: absolute net_amount z-score >= 2.0 within category",
        "Refund history flag: user has >= 1 prior refunded order",
    ]
    y = draw_bullets(c, x, y, scoring_bullets)

    y -= 4
    y = draw_paragraph(c, x, y, "Signal weights:", font_name="Helvetica-Bold", font_size=10)
    weight_lines = [f"{k} = {v}" for k, v in WEIGHTS.items()]
    y = draw_bullets(c, x, y, weight_lines, font_size=9)

    y -= 4
    y = draw_paragraph(c, x, y, "Risk band thresholds:", font_name="Helvetica-Bold", font_size=10)
    y = draw_bullets(c, x, y, RISK_BAND_RULE, font_size=9)
    y -= 8

    # -----------------------------
    # 5. Investigation queue + examples
    # -----------------------------
    y = draw_heading(c, x, y, "5. Investigation Queue + Examples", level=2)
    y = draw_paragraph(
        c, x, y,
        "The queue ranks orders by descending risk_score and then by descending net_amount, so operations teams review the most suspicious "
        "and potentially highest-value cases first."
    )

    queue_bullets = [
        "Priority fields included: rank, order_id, user_id, risk_score, risk_band, top 3 reasons, recommended action, amount, pincode.",
        "Recommended actions include ALLOW, MONITOR, SOFT_FRICTION_OTP, CALL_VERIFICATION, OTP_ADDRESS_CONFIRMATION, and MANUAL_REVIEW.",
    ]
    y = draw_bullets(c, x, y, queue_bullets)

    if not top_queue.empty:
        example_lines = []
        for _, r in top_queue.iterrows():
            example_lines.append(
                f"Rank {int(r['rank'])}: order {r['order_id']} | score={r['risk_score']} | band={r['risk_band']} | "
                f"reasons=({r['reason_1']}, {r['reason_2']}, {r['reason_3']}) | action={r['recommended_action']}"
            )
        y = draw_paragraph(c, x, y, "Top queue examples:", font_name="Helvetica-Bold", font_size=10)
        y = draw_bullets(c, x, y, example_lines, font_size=9)

    y -= 8

    # -----------------------------
    # 6. Controls + experiment plan
    # -----------------------------
    y = draw_heading(c, x, y, "6. Controls + Experiment Plan", level=2)
    controls_bullets = [
        "Coupon abuse control: restrict coupon redemption by device_id and shipping_pincode when repeat abuse is detected.",
        "Friction control: apply OTP / step-up verification for medium-risk new-user coupon orders.",
        "COD control: add OTP address confirmation for COD orders in high-RTO pincodes.",
        "Manual review control: send highest-risk orders to analyst queue before fulfillment.",
        "Payment control: trigger call verification or alternate payment flow after repeated failed attempts.",
    ]
    y = draw_bullets(c, x, y, controls_bullets)

    experiment_bullets = [
        "Run a 2–4 week pilot comparing treated vs untreated traffic for selected controls.",
        "Primary success metrics: refund_amount, refund_rate, RTO rate, high-risk order share, and investigator hit rate.",
        "Secondary guardrails: conversion rate, payment success rate, average order value, and customer support contacts.",
        "Measure success weekly using kpi_weekly and queue outcomes before expanding controls platform-wide.",
    ]
    y = draw_paragraph(c, x, y, "Experiment plan:", font_name="Helvetica-Bold", font_size=10)
    y = draw_bullets(c, x, y, experiment_bullets)
    y -= 8

    # -----------------------------
    # 7. 30-day impact estimate + risks/limitations
    # -----------------------------
    y = draw_heading(c, x, y, "7. 30-Day Impact Estimate + Risks / Limitations", level=2)

    impact_bullets = [
        f"Current refund-based monthly loss proxy: {fmt_num(monthly_loss_proxy, 2)}.",
        f"Conservative scenario: if 20% of this loss is preventable and controls capture 50% of preventable loss, estimated 30-day savings are {fmt_num(estimated_30_day_savings, 2)}.",
        "This estimate is directional and should be validated with pilot results and true fraud-confirmation labels.",
    ]
    y = draw_paragraph(c, x, y, "30-day impact estimate:", font_name="Helvetica-Bold", font_size=10)
    y = draw_bullets(c, x, y, impact_bullets)

    limitation_bullets = [
        "Refund amount and RTO are proxies, not perfect fraud labels.",
        "Current rule set may understate risk if key fields such as device_id, shipment outcomes, or chargeback signals are sparse.",
        "Thresholds are heuristic and should be recalibrated once analyst feedback and confirmed case labels are available.",
        "The current snapshot is rule-based; future versions should add supervised / anomaly models and closed-loop feedback from investigators.",
    ]
    y = draw_paragraph(c, x, y, "Risks / limitations:", font_name="Helvetica-Bold", font_size=10)
    y = draw_bullets(c, x, y, limitation_bullets)

    # footer-like close
    y -= 6
    y = draw_paragraph(
        c, x, y,
        "Conclusion: the pipeline already produces the core monitoring assets needed for an always-on fraud dashboard — curated data, explainable scoring, weekly KPIs, pattern summaries, and a ranked investigation queue."
    )

    c.save()
    print(f"✅ Wrote: {out_path}")

# =========================
# Validations
# =========================
def validate_fact(fact: pd.DataFrame) -> None:
    if "order_id" not in fact.columns:
        raise ValueError("fact_orders_enriched must have order_id")
    dup = fact["order_id"].duplicated().sum()
    if dup > 0:
        raise ValueError(f"Join explosion: {dup} duplicate order_id rows in fact_orders_enriched")
    if (fact["risk_score"] < 0).any() or (fact["risk_score"] > 100).any():
        raise ValueError("risk_score out of bounds 0..100")


# =========================
# Main pipeline
# =========================
def main():
    print("📦 Loading raw files from:", RAW_DIR)

    users = read_csv("users.csv")
    sessions = read_csv("sessions.csv")
    orders = read_csv("orders.csv")
    order_items = read_csv("order_items.csv")
    payments = read_csv("payments.csv")
    shipments = read_csv("shipments.csv")
    refunds = read_csv("refunds.csv")
    coupons = read_csv("coupons.csv")
    products_json = read_json("products.json")
    products = flatten_products(products_json)

    fact = build_fact_orders_enriched(
        users=users,
        sessions=sessions,
        orders=orders,
        order_items=order_items,
        payments=payments,
        shipments=shipments,
        refunds=refunds,
        coupons=coupons,
        products=products,
    )

    fact = compute_score_and_reasons(fact)
    weekly = build_user_weekly(fact)
    queue = build_queue(fact)

    validate_fact(fact)

    # 1) Core ETL outputs
    write_csv(fact, DATA_DIR / "fact_orders_enriched.csv")
    write_csv(weekly, DATA_DIR / "fact_user_risk_weekly.csv")
    write_csv(queue, DATA_DIR / "investigation_queue.csv")

    # 2) Analysis outputs
    artifacts = build_analysis_artifacts(fact, weekly, queue)
    kpi_weekly = artifacts["kpi_weekly"]
    patterns_summary = artifacts["patterns_summary"]

    write_csv(kpi_weekly, ANALYSIS_DIR / "kpi_weekly.csv")
    write_csv(patterns_summary, ANALYSIS_DIR / "patterns_summary.csv")
    export_analysis_report_html(kpi_weekly, patterns_summary, ANALYSIS_DIR / "analysis_report.html")
    # 2b) Part C analytics outputs
    part_c = build_part_c_outputs(fact, kpi_weekly)

    for name, df_out in part_c.items():
        write_csv(df_out, ANALYSIS_DIR / f"{name}.csv")

        segment_outputs = {k: v for k, v in part_c.items() if "top_loss" in k or "top_risk" in k}

        export_analysis_report_html(
        kpi_weekly=kpi_weekly,
        patterns_summary=patterns_summary,
        out_path=ANALYSIS_DIR / "analysis_report.html",
        weekly_diagnosis=part_c.get("weekly_diagnosis"),
        named_patterns=part_c.get("named_patterns"),
        investigation_table=part_c.get("investigation_table_top2_patterns"),
        segment_outputs=segment_outputs,
    )
    # 3) Dashboard outputs
    export_charts(kpi_weekly, fact)
    export_dashboard_xlsx(fact, weekly, queue, kpi_weekly, patterns_summary)

    # 4) Final story outputs
    export_final_memo_pdf(
        fact=fact,
        kpi_weekly=kpi_weekly,
        patterns_summary=patterns_summary,
        queue=queue,
        out_path=FINAL_STORY_DIR / "final_memo.pdf"
    )

    print("\n🎉 All done.")
    print("Core CSVs in /data")
    print("Analysis report in /analysis")
    print("Dashboard in /dashboard")
    print("Final memo in /final_story")


if __name__ == "__main__":
    main()