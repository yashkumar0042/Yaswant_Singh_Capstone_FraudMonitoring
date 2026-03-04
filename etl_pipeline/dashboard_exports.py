from __future__ import annotations

import pandas as pd
import matplotlib.pyplot as plt

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from .paths import DASHBOARD_DIR, EXPORTS_DIR


def export_charts(kpi_weekly: pd.DataFrame, fact: pd.DataFrame) -> None:
    if not kpi_weekly.empty and "week_start" in kpi_weekly.columns:
        plt.figure()
        plt.plot(kpi_weekly["week_start"], kpi_weekly["refund_amount"])
        plt.xticks(rotation=45, ha="right")
        plt.title("Weekly Refund Amount Trend")
        plt.tight_layout()
        path1 = EXPORTS_DIR / "weekly_refunds_trend.png"
        plt.savefig(path1, dpi=160)
        plt.close()
        print(f"✅ Wrote: {path1}")

    if "risk_band" in fact.columns:
        band_counts = fact["risk_band"].value_counts(dropna=False)
        plt.figure()
        plt.pie(band_counts.values, labels=band_counts.index.astype(str), autopct="%1.1f%%")
        plt.title("Risk Band Split (All Orders)")
        plt.tight_layout()
        path2 = EXPORTS_DIR / "risk_band_split.png"
        plt.savefig(path2, dpi=160)
        plt.close()
        print(f"✅ Wrote: {path2}")


def autosize_excel_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(45, max(10, max_len + 2))


def export_dashboard_xlsx(
    fact: pd.DataFrame,
    weekly: pd.DataFrame,
    queue: pd.DataFrame,
    kpi_weekly: pd.DataFrame,
    patterns_summary: pd.DataFrame
) -> None:
    xlsx_path = DASHBOARD_DIR / "dashboard.xlsx"

    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        kpi_weekly.to_excel(writer, sheet_name="KPI_Weekly", index=False)
        patterns_summary.to_excel(writer, sheet_name="Patterns", index=False)
        queue.head(500).to_excel(writer, sheet_name="Investigation_Queue", index=False)
        fact.head(2000).to_excel(writer, sheet_name="Sample_Fact", index=False)

    wb = load_workbook(xlsx_path)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        autosize_excel_columns(ws)

    wb.save(xlsx_path)
    print(f"✅ Wrote: {xlsx_path}")